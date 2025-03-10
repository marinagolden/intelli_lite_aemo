"""
Created on 25/10/2021
"""
__author__    = "Mostafa Naemi"
__copyright__ = "Cornwall Insight Australia"

import numpy     as np
import pandas    as pd
import utilities as ut
import openpyxl  as pyxl
from openpyxl import load_workbook,Workbook

import subprocess
import sys

def install(xlwings):
    subprocess.check_call([sys.executable, "-m", "pip", "install", xlwings])

import xlwings
import os
from   pathlib          import Path
    ####################### POST PROCESSING FUNCTION ############################


# ==================================================================================================================== 
def post_process_battery_results(model_outputs,num_cycle_cum,input,battery_current_cap,date,resolution=30):
    """ this post process the output of the model and prepare that for the final post 
    process and also next solve loops
    -----
    model_output: output of the LP model
    output      : output of the this function for previous intervals
    input       : input data selected based on scenario
    date        : the date specified in the loop in main code script
    resolution  : resolution of the model
    """
    scenario_manager = input['Scn_mgr']
    disch_efficiency = scenario_manager['Disch_efncy'].iloc[0]         
    resolution       = resolution / 60
    vars             = model_outputs['Variables']
    percent_method   = scenario_manager['PPA load curve'].iloc[0] == 'None'
    ppa_sim          = scenario_manager['PPA load'      ].iloc[0] == 'Yes'
    load_obligation = scenario_manager ['PPA Load Obligation'].iloc[0] == 'Strict'

    vars_list = [] 
    for index in range(len(vars)):
        vars_list.append([vars[index].name , vars[index].varValue])
    
    #create a df from decision vars and their optimal values
    df             = pd.DataFrame(vars_list,columns  = ['Variable' , 'Value'])
    df['Period'  ] = df['Variable'].str.split('_').str[-1]
    df['Variable'] = df.apply(lambda row : row['Variable'].replace(str('_'+row['Period']), ''), axis=1)
        
    # df = ut.postproc_period_to_datetime(df,date) 
    df_pivot = df.pivot_table(columns = 'Variable',index=['Period'],values ='Value')
    
    if ppa_sim and percent_method:   
        df_pivot['PPA_Load_Exceed']=0 

    vars_throughput = { 'BTM_co_loc_gen_energy_export'          :1, 
                        'BTM_energy_vol_charge_curtail'         :1,
                        'BTM_energy_vol_charge_export_avail_gen':1, 
                        'BTM_lreg_vol_charge_export_avail_gen'  :scenario_manager['Svc_throuput_Lreg'].iloc[0],
                        'FCAS_l5min_volumes_charge'             :scenario_manager['Svc_throuput_L5'  ].iloc[0], 
                        'FCAS_l60sec_volumes_charge'            :scenario_manager['Svc_throuput_L60' ].iloc[0],
                        'FCAS_l6sec_volumes_charge'             :scenario_manager['Svc_throuput_L6'  ].iloc[0],
                        'FCAS_l1sec_volumes_charge'             :scenario_manager['Svc_throuput_L1'  ].iloc[0],
                        'FCAS_lreg_volumes_charge'              :scenario_manager['Svc_throuput_Lreg'].iloc[0], 
                        'FCAS_r5min_volumes_discharge'          :scenario_manager['Svc_throuput_R5'  ].iloc[0],
                        'FCAS_r60sec_volumes_discharge'         :scenario_manager['Svc_throuput_R60' ].iloc[0],
                        'FCAS_r6sec_volumes_discharge'          :scenario_manager['Svc_throuput_R6'  ].iloc[0],
                        'FCAS_r1sec_volumes_discharge'          :scenario_manager['Svc_throuput_R1'  ].iloc[0],
                        'FCAS_rreg_volumes_discharge'           :scenario_manager['Svc_throuput_Rreg'].iloc[0],
                        'energy_volumes_charge'                 :1, 
                        'BTM_Grid_to_load'                      :1,
                        'BTM_BESS_to_load'                      :1,
                        'BTM_solar_to_load'                     :1,
                        'energy_volumes_discharge'              :1, 
                        'PPA_offtake'                           :1, 
                        'PPA_Load_Exceed'                       :1}
    raise_svc = ['FCAS_r5min_volumes_discharge','FCAS_r60sec_volumes_discharge','FCAS_r6sec_volumes_discharge','FCAS_r1sec_volumes_discharge','FCAS_rreg_volumes_discharge','energy_volumes_discharge','BTM_BESS_to_load']
      
    df_pivot['battery_current_cap'] = battery_current_cap
    hdr                             = pd.MultiIndex.from_product([['Value'],list(df_pivot.columns)])
    df_pivot.columns                = hdr
    for k in vars_throughput.keys():
        if k in  df_pivot['Value'].columns:
            df_pivot[('Throughput',k)] = df_pivot[('Value',k)] * vars_throughput[k] * resolution

    df_pivot.reset_index(inplace=True)
    df_pivot['Period'] = df_pivot['Period'].astype(int)
    df_pivot           = df_pivot.sort_values('Period')
    df_pivot           = df_pivot.head(int(24/resolution))

    num_cycles         = (df_pivot['Throughput'][raise_svc].sum().sum())/(battery_current_cap)/disch_efficiency # 1 full cycle : (1 full discharge + 1 full charge)
    num_cycle_cum      = num_cycle_cum + num_cycles   
    end_of_loop_soc    = df_pivot.loc[df_pivot.Period == 24 / resolution,('Value','soc_end_period')].iloc[0] ## should be 48 in square brackets but using work around
    post_proc_dict     = {'df':df_pivot , 'end_of_loop_soc': end_of_loop_soc, 'num_cycles':num_cycle_cum}
    return post_proc_dict

# ==================================================================================================================== 
def revenue_calc_postproc(df,input,resolution=30):
    """this function calculates the revenus and costs and add them as separate colunms to 30min data
    Inputs:
    df    : output of the post_process_results function (30min data)
    input : dictionary of all inputs (such as scenario manager, DLF/MLF, DNSP etc.)
    """
    #output is the dataframe of the post process results (above func)
    # input all data import 
    price       = input['Price'      ] #price scn
    scn_mgr     = input['Scn_mgr'    ]
    DLF         = input['DLF'        ]
    DLF_cogen   = input['DLF_cogen'  ]
    MLF         = input['MLF'        ]
    DNSP        = input['DNSP_tariff']
    DNSP        = DNSP[['Datetime','DNSP_volume_tariff']]
    LGC_STC     = input['LGC_STC'    ]
    LReg_thrput = scn_mgr['Svc_throuput_Lreg'].iloc[0]
    RReg_thrput = scn_mgr['Svc_throuput_Rreg'].iloc[0]
    chrg_efficiency         = scn_mgr['Chrg_efncy'       ].iloc[0] 
    disch_efficiency        = scn_mgr['Disch_efncy'      ].iloc[0] 

    resolution          = resolution / 60
    retail_margin_rev   = 1-scn_mgr['retail_margin_rev'].iloc[0]
    retail_margin       = 1+scn_mgr['retail_margin_rev'].iloc[0]

    front_meter = scn_mgr['Behind_meter'].iloc[0] == 'No'
    BTM = not front_meter

    var_price_dict = {  'BTM_co_loc_gen_energy_export'          :scn_mgr['Location'].iloc[0], 
                        'BTM_energy_vol_charge_curtail'         :scn_mgr['Location'].iloc[0],
                        'BTM_energy_vol_charge_export_avail_gen':scn_mgr['Location'].iloc[0],
                        'BTM_lreg_vol_charge_export_avail_gen'  :'LOWERREGRRP',
                        'FCAS_l5min_volumes_charge'             :'LOWER5MINRRP',   
                        'FCAS_l60sec_volumes_charge'            :'LOWER60SECRRP', 
                        'FCAS_l6sec_volumes_charge'             :'LOWER6SECRRP',
                        'FCAS_l1sec_volumes_charge'             :'LOWER1SECRRP',
                        'FCAS_lreg_volumes_charge'              :'LOWERREGRRP', 
                        'FCAS_r5min_volumes_discharge'          :'RAISE5MINRRP',
                        'FCAS_r60sec_volumes_discharge'         :'RAISE60SECRRP', 
                        'FCAS_r6sec_volumes_discharge'          :'RAISE6SECRRP',
                        'FCAS_r1sec_volumes_discharge'          :'RAISE1SECRRP',
                        'FCAS_rreg_volumes_discharge'           :'RAISEREGRRP', 
                        'energy_volumes_charge'                 :scn_mgr['Location'].iloc[0],
                        'energy_volumes_discharge'              :scn_mgr['Location'].iloc[0],
                        'BTM_BESS_to_load'                      :scn_mgr['Location'].iloc[0], 
                        'BTM_solar_to_load'                     :scn_mgr['Location'].iloc[0], 
                        'BTM_Grid_to_load'                      :scn_mgr['Location'].iloc[0],
                        }
    
    fcas_list = [   'BTM_lreg_vol_charge_export_avail_gen'  ,
                    'FCAS_l5min_volumes_charge'             ,
                    'FCAS_l60sec_volumes_charge'            , 
                    'FCAS_l6sec_volumes_charge'             ,
                    'FCAS_l1sec_volumes_charge'             ,
                    'FCAS_lreg_volumes_charge'              , 
                    'FCAS_r5min_volumes_discharge'          ,
                    'FCAS_r60sec_volumes_discharge'         , 
                    'FCAS_r6sec_volumes_discharge'          ,
                    'FCAS_r1sec_volumes_discharge'          ,
                    'FCAS_rreg_volumes_discharge']
    
    network_cost_list = ['energy_volumes_charge'   ,
                         'FCAS_lreg_volumes_charge',
                         'BTM_Grid_to_load']

    LGC_cost_list = ['energy_volumes_charge'   ,
                     'FCAS_lreg_volumes_charge']

    for k in var_price_dict.keys():
        df[('Rev',k)] = df[('Value',k)].abs() * resolution * price[var_price_dict[k]].values 
    df[('Rev','energy_volumes_charge'       )] = - retail_margin     * df[('Rev','energy_volumes_charge'       )] * MLF['MLF_load'      ].iloc[0] * DLF
    df[('Rev','BTM_Grid_to_load'            )] = - retail_margin     * df[('Rev','BTM_Grid_to_load')] * MLF['MLF_load'                       ].iloc[0] * DLF 
    df[('Rev','energy_volumes_discharge'    )] =   retail_margin_rev * df[('Rev','energy_volumes_discharge'    )] * MLF['MLF_generation'].iloc[0] * DLF
    df[('Rev','BTM_co_loc_gen_energy_export')] =                       df[('Rev','BTM_co_loc_gen_energy_export')] * MLF['co gen'        ].iloc[0] * DLF_cogen
    

    if BTM:
        df[('Rev','LGC'                         )] =  (  
               df[('Value','energy_volumes_charge'       )] *resolution *LGC_STC['LGC price'].iloc[0] *LGC_STC['LGC percentage'].iloc[0] * DLF * MLF['MLF_load'      ].iloc[0]
            + df[('Value','energy_volumes_discharge'    )] *resolution *LGC_STC['LGC price'].iloc[0] *LGC_STC['LGC percentage'].iloc[0] * DLF * MLF['co gen'].iloc[0]
            #+ df[('Value','BTM_lreg_vol_charge_export_avail_gen'    )] *resolution *LGC_STC['LGC price'].iloc[0] *LGC_STC['LGC percentage'].iloc[0] * DLF * MLF['MLF_load'].iloc[0] * LReg_thrput*chrg_efficiency*disch_efficiency
            + df[('Value','FCAS_rreg_volumes_discharge' )] *resolution *LGC_STC['LGC price'].iloc[0] *LGC_STC['LGC percentage'].iloc[0] * DLF * MLF['co gen'].iloc[0] * RReg_thrput
            #+ df[('Value','BTM_lreg_vol_charge_export_avail_gen')] *resolution *-LGC_STC['LGC price'].iloc[0] * DLF * MLF['co gen'].iloc[0] * LReg_thrput*chrg_efficiency*disch_efficiency
            #+ df[('Value','BTM_co_loc_gen_energy_export')] *resolution *LGC_STC['LGC price'].iloc[0] * DLF * MLF['co gen'].iloc[0]
            + df[('Value','BTM_co_loc_gen_energy_export')] *resolution *LGC_STC['LGC price'].iloc[0]*LGC_STC['LGC percentage'].iloc[0] * DLF * MLF['co gen'].iloc[0]
            #+ df[('Value','BTM_energy_vol_charge_curtail')] *resolution *-LGC_STC['LGC price'].iloc[0] * DLF * MLF['co gen'].iloc[0]*chrg_efficiency*disch_efficiency # Take away RTE losses
            #+ df[('Value','BTM_energy_vol_charge_export_avail_gen')] *resolution *-LGC_STC['LGC price'].iloc[0]* DLF * MLF['co gen'].iloc[0]*chrg_efficiency*disch_efficiency 
                                                      )

    for s in fcas_list:
        df[('FCAS_Energy_Revenue',s)] =  df[('Throughput',s)] * price[scn_mgr['Location'].iloc[0]].values
    
    df[('FCAS_Energy_Revenue','FCAS_rreg_volumes_discharge')] =  df[('FCAS_Energy_Revenue','FCAS_rreg_volumes_discharge')] * MLF['MLF_generation'].iloc[0] * DLF
    df[('FCAS_Energy_Revenue','FCAS_lreg_volumes_charge'   )] =  df[('FCAS_Energy_Revenue','FCAS_lreg_volumes_charge'   )] * MLF['MLF_load'      ].iloc[0] * DLF
    
    df[('Cost','Network_cost')] = df['Throughput'][network_cost_list].sum(axis=1)  *  DNSP['DNSP_volume_tariff'].values* MLF['MLF_load'].iloc[0] * DLF
    if BTM:
         df[('Cost','LGC_STC_cost')] = 0
    else:
        df[('Cost','LGC_STC_cost')] = df['Throughput'][LGC_cost_list    ].sum(axis=1)  * (LGC_STC['LGC price'].iloc[0]*LGC_STC['LGC percentage'].iloc[0]+LGC_STC['STC price'].iloc[0]*LGC_STC['STC percentage'].iloc[0])*\
        (1-scn_mgr['Chrg_efncy'].iloc[0])* DLF*  MLF['MLF_load'].iloc[0]
    
    if 'price_diff' in input.keys():
        ppa_price_diff = input  ['price_diff'    ]
        percent_method = scn_mgr['PPA load curve'].iloc[0] == 'None'
        PPA_price      = input  ['PPA_price'     ]

        df[('Rev','PPA_offtake')   ] =  DLF_cogen *MLF['co gen'].iloc[0]* df[('Value','PPA_offtake')].abs() * resolution * ppa_price_diff['Price_Diff'].values 
        df['Price','PPA Price Diff'] =  ppa_price_diff['Price_Diff'].values *  df[('Value','PPA_offtake')]

        if not percent_method:
            df[('Rev','PPA_offtake')] =  df[('Rev','PPA_offtake')] - df[('Value','PPA_Load_Exceed')].abs() * resolution * (PPA_price['PPA_Price'].values -  ppa_price_diff['Price_Diff'].values)
        cols = ['Period','Value' , 'Throughput' , 'Rev' ,'FCAS_Energy_Revenue', 'Cost']
        df = df[cols]
    return df

# ==================================================================================================================== 
def avg_running_cost_calc(df, running_30days_cost, num_days =30):
    charge_list = [ 'BTM_co_loc_gen_energy_export'          , 
                    'BTM_energy_vol_charge_curtail'         ,
                    'BTM_energy_vol_charge_export_avail_gen', 
                    'FCAS_l5min_volumes_charge'             ,
                    'FCAS_l60sec_volumes_charge'            , 
                    'FCAS_l6sec_volumes_charge'             ,
                    'FCAS_l1sec_volumes_charge'             ,
                    'FCAS_lreg_volumes_charge'              ,
                    'energy_volumes_charge']

    discharge_list = [  'FCAS_r5min_volumes_discharge'  ,
                        'FCAS_r60sec_volumes_discharge' , 
                        'FCAS_r6sec_volumes_discharge'  ,
                        'FCAS_r1sec_volumes_discharge'  ,
                        'FCAS_rreg_volumes_discharge'   ,
                        'energy_volumes_discharge',
                        'BTM_Grid_to_load',
                        'BTM_BESS_to_load' ,
                        'BTM_solar_to_load',
                         ]
    
    cost             = df['Cost'                                            ].sum().sum()
    energy_chrg_cost = df[('Rev','energy_volumes_charge')                   ].sum().sum()
    lreg_chrg_cost   = df[('FCAS_Energy_Revenue','FCAS_lreg_volumes_charge')].sum().sum()
    throughput       = df['Throughput'][discharge_list                      ].sum().sum()

    cost_per_MWh_disch =  -( cost +energy_chrg_cost+lreg_chrg_cost) / (throughput+.0000001) 
    running_30days_cost.append(cost_per_MWh_disch)
    l = len(running_30days_cost)
    if  l > 30:
        avg_running_30days_avg_cost = sum(running_30days_cost[-num_days:])/num_days
        # drop old days costs
        running_30days_cost = running_30days_cost[-num_days:]
    else: 
        avg_running_30days_avg_cost = sum(running_30days_cost[-num_days:])/l  
    return  running_30days_cost,avg_running_30days_avg_cost

# ==================================================================================================================== 
def daily_spread_calc(df,avg_today_spread,running_30days_cost,min_avg_spread,max_avg_spread, rreg_min_avg_spread,rreg_max_avg_spread):
    if df.empty:
        running_30days_cost         = [0]
        avg_running_30days_avg_cost =  0
    else:
        running_30days_cost,avg_running_30days_avg_cost = avg_running_cost_calc(df,running_30days_cost)


    if avg_today_spread > min_avg_spread + min(50,avg_running_30days_avg_cost):
        daily_spread    = max_avg_spread
    else:
        daily_spread    = min_avg_spread + min(50,avg_running_30days_avg_cost)

    if avg_today_spread   > rreg_min_avg_spread + avg_running_30days_avg_cost:
        rreg_daily_spread =  rreg_max_avg_spread 
    else:
        rreg_daily_spread = rreg_min_avg_spread + avg_running_30days_avg_cost

    return daily_spread,rreg_daily_spread,running_30days_cost

# ==================================================================================================================== 
def combine_dfs_dict(dfs_dict, resolution):
    df =pd.DataFrame()
    for key in dfs_dict:
        df_ = dfs_dict[key]
        df_ = ut.postproc_period_to_datetime(df_,key, duration = resolution)
        df  = pd.concat([df,df_],axis=0)
    return df

# ==================================================================================================================== 
def add_price_to_df_final(df_final,price_scn):
    price_scn . drop('PERIODID',axis=1,inplace=True)
    price_scn . set_index('Datetime',inplace=True)
    hdr       = pd.MultiIndex.from_product([['Price'],list(price_scn.columns)])
    price_scn . columns = hdr
    df_final  . set_index('Datetime',inplace=True)
    df_final  = df_final.merge(price_scn,left_index=True,right_index = True)
    df_final  . reset_index(inplace=True)
    return df_final

# ==================================================================================================================== 
def make_summary(data, scenario , scn_mgr, BTM, curtail,residential_load,Res_Load, aemo_fees, MLF, DLF, DNSP, LGC, region,resolution,out_dir,ppa_sim,PPA_price,Solve_status_dict):
    """this function creates the final monthly summary csv required for client version"""
    resolution = resolution / 60 #convert to hour

    # Add year,date,day,month to dataframe
    data[('Datetime','Year' ) ] = data[('Datetime','')].dt.year
    data[('Datetime','Month') ] = data[('Datetime','')].dt.month
    data[('Datetime','Day'  ) ] = data[('Datetime','')].dt.day
    data[('Datetime','Date' ) ] = data[('Datetime','')].dt.date
    data['Price'              ] = data['Price'        ].astype(float)
    # Add curtailment profile to dataframe for BTM cases
    if BTM:
        data = merge_curtail(data,curtail)
        if not residential_load.empty:
            data = merge_curtail(data,residential_load)
        
    svc_list =  [   'BTM_lreg_vol_charge_export_avail_gen'  ,
                    'FCAS_l5min_volumes_charge'             ,
                    'FCAS_l1sec_volumes_charge'             ,
                    'FCAS_l6sec_volumes_charge'             ,
                    'FCAS_l60sec_volumes_charge'            ,
                    'FCAS_lreg_volumes_charge'              ,
                    'FCAS_r5min_volumes_discharge'          ,
                    'FCAS_r60sec_volumes_discharge'         ,
                    'FCAS_r1sec_volumes_discharge'          ,
                    'FCAS_r6sec_volumes_discharge'          ,
                    'FCAS_rreg_volumes_discharge'           ,
                    'energy_volumes_discharge'              ,
                    'energy_volumes_charge'                 ,
                    'Total_Lreg'                            ]

    rev_list =  [   'BTM_lreg_vol_charge_export_avail_gen',
                    'FCAS_l5min_volumes_charge'           ,
                    'FCAS_l1sec_volumes_charge'           ,
                    'FCAS_l6sec_volumes_charge'           ,
                    'FCAS_l60sec_volumes_charge'          ,
                    'FCAS_lreg_volumes_charge'            ,
                    'FCAS_r5min_volumes_discharge'        ,
                    'FCAS_r60sec_volumes_discharge'       ,
                    'FCAS_r1sec_volumes_discharge'        ,
                    'FCAS_r6sec_volumes_discharge'        ,
                    'FCAS_rreg_volumes_discharge'         ,
                    'energy_volumes_discharge'            ]

    svc_raise = [   'energy_volumes_discharge'      ,
                    'FCAS_rreg_volumes_discharge'   ,
                    'FCAS_r1sec_volumes_discharge'  ,
                    'FCAS_r6sec_volumes_discharge'  ,
                    'FCAS_r60sec_volumes_discharge' ,
                    'FCAS_r5min_volumes_discharge'  ]

    lreg_rev  = ['BTM_lreg_vol_charge_export_avail_gen',
                 'FCAS_lreg_volumes_charge'            ]
    # fcas_rev_cost = ['FCAS_lreg_volumes_charge']
    # cert_cost = ['LGC_STC_cost']
    # energy_cost = ['energy_volumes_charge'] 
    data = merge_DLF_MLF(data,MLF,DLF)
    data = merge_DNSP   (data,DNSP   )
    data = merge_LGC    (data,LGC    )

    if ppa_sim:
        data = merge_PPA(data,PPA_price)
        data[('Rev' , 'PPA offtake Spot' )] = data['Throughput','PPA_offtake'] * data['Price', region] * data['MLF','co gen'] * data['DLF','DLF_cogen']
        data[('Rev' , 'PPA offtake Total')] = data['Throughput','PPA_offtake'] * data['Price', 'PPA' ] * data['MLF','co gen'] * data['DLF','DLF_cogen']

        data[('Throughput' , 'PPA offtake Solar'  )] = np.where(data['Value','BTM_co_loc_gen_energy_export'] >= data['Value' , 'PPA_offtake'],data['Throughput' , 'PPA_offtake'],data['Throughput','BTM_co_loc_gen_energy_export'] )
        data[('Throughput' , 'PPA Load Satisfied' )] = data['Throughput','PPA_offtake'           ] - data['Throughput','PPA_Load_Exceed']
        data[('Cost' , 'PPA Load Exceed'          )] = data['Throughput','PPA_Load_Exceed'       ] * data['Price', region] 
        data[('Rev' , 'PPA Load Satisfied Spot'   )] = data[('Throughput' , 'PPA Load Satisfied')] * data['Price', region] * data['MLF','co gen'] * data['DLF','DLF_cogen']
        data[('Rev' , 'PPA Revenue Load Satisfied')] = data[('Throughput' , 'PPA Load Satisfied')] * data['Price', 'PPA' ] * data['MLF','co gen'] * data['DLF','DLF_cogen']

    data[('Throughput' , 'Period'  )] = data['Throughput' ][svc_raise].sum(axis=1)
    data[('Value','Total_Lreg'     )] = data['Value'      ][lreg_rev ].sum(axis=1)
    data[('Throughput','Total_Lreg')] = data['Throughput' ][lreg_rev ].sum(axis=1)

    if BTM:
        data[('Rev', 'BTM_co_loc_gen_energy_export'                       )] = (data[('Value', 'BTM_co_loc_gen_energy_export')] + data[('Value', 'BTM_lreg_vol_charge_export_avail_gen')]*scn_mgr['Svc_throuput_Lreg'].iloc[0])  * resolution * data['Price'].iloc[:,0] * data['MLF','co gen'] * data['DLF','DLF_cogen']
        data[('Rev', 'BTM_energy_vol_charge_export_avail_gen'             )] =  data[('Value', 'BTM_energy_vol_charge_export_avail_gen')] * resolution * data['Price'].iloc[:,0] * data['MLF','co gen'] * data['DLF','DLF_cogen']
        data[('FCAS_Energy_Revenue','BTM_lreg_vol_charge_export_avail_gen')] =  data[('FCAS_Energy_Revenue','BTM_lreg_vol_charge_export_avail_gen')] * data['MLF','co gen'] * data['DLF','DLF_cogen']

        #curtailment for reserving connection line for bess fcas provision
        ix = data['Price',region]>=-data['LGC','Price']
        data    ['Value','BESS_Con_Res_Curtail'       ] = 0
        data.loc[ix,[('Value','BESS_Con_Res_Curtail')]] = (data.loc[ix,[('Curtail','net_co_locate_gen_avail_export')]].values -data.loc[ix,[('Value','BTM_co_loc_gen_energy_export')]].values +  data.loc[ix,[('Value','BTM_energy_vol_charge_export_avail_gen')]].values) * resolution
        data    [('Cost','BESS_Con_Res_Curtail'      )] = data['Value','BESS_Con_Res_Curtail'] * data['Price',region]     
    
    aemo_costs   = aemo_cost            (aemo_fees,region)
    data_monthly,data_sum = monthly_data(data,svc_list)
    data_monthly = market_profit        (data_monthly,aemo_costs,DNSP,rev_list)
    data_monthly = revenue_brkdwn       (data_monthly,lreg_rev,ppa_sim)
    data_monthly = operation_calc       (data_monthly)
    data_monthly = enabelement_calc     (data_monthly)
    data_monthly = throughput_calc      (data_monthly,BTM,resolution,Res_Load)
    data_monthly = throughput_perc_calc (data_monthly)
    if BTM:
        data_monthly = BTM_calc(data_monthly,resolution,Res_Load)
        if ppa_sim:
            data_monthly = PPA_calc(data_monthly)
    
    data_monthly = rev_per_MWh(data_monthly,data_sum)
    data_monthly = DLF_MLF_rev(data_monthly         )
    export_summary(data_monthly,BTM,Res_Load,ppa_sim,out_dir , scenario,Solve_status_dict)

# ==================================================================================================================== 
def market_profit(data_gp,aemo_costs,DNSP,rev_list):
    network_cost         = ['Network_cost']
    aemo_customer_cost   = aemo_costs['aemo_customer']
    aemo_generator_cost  = aemo_costs['aemo_gen'     ]  
    #%% storage summary
    Market_profit_list = [  'Energy and FCAS Revenue'         ,
                            'Enviromental Certificate Costs'  ,
                            'Energy Costs','Lreg Energy Costs',
                            'Grid to Load Costs'              ,
                            'AEMO Costs'                      ,
                            'Network Costs (variable)'        ,
                            'Network Costs (fixed)'           ]
    fcas_rev_energy    = ['FCAS_rreg_volumes_discharge'       ]

    data_gp['calcs_summary','Energy and FCAS Revenue'       ] =   data_gp['Rev' ][rev_list].sum(axis=1) + data_gp['FCAS_Energy_Revenue'][fcas_rev_energy].sum(axis=1)
    data_gp['calcs_summary','Enviromental Certificate Costs'] =   data_gp['Cost']['LGC_STC_cost']
    data_gp['calcs_summary','Energy Costs'                  ] =   data_gp['Rev' ]['energy_volumes_charge']
    data_gp['calcs_summary','Grid to Load Costs'            ] =   data_gp['Rev' ]['BTM_Grid_to_load']
    data_gp['calcs_summary','Lreg Energy Costs'             ] =   data_gp['FCAS_Energy_Revenue']['FCAS_lreg_volumes_charge']
    data_gp['calcs_summary','AEMO Costs'                    ] = -(data_gp['Throughput','energy_volumes_charge'] + data_gp['Throughput','FCAS_lreg_volumes_charge']) * aemo_customer_cost  \
        + (data_gp['Throughput','energy_volumes_discharge'] + data_gp['Throughput','FCAS_rreg_volumes_discharge']) * aemo_generator_cost 
    data_gp['calcs_summary','Network Costs (variable)'      ] =   data_gp['Cost'][network_cost].sum(axis=1)
    data_gp['calcs_summary','Network Costs (fixed)'         ] = - data_gp[('Datetime','Day')] * data_gp['DNSP','Fixed ($/day)']
    data_gp['calcs_summary','Network Cost (fixed) indicative site specific (TBC with DNSP)'] = 0
    data_gp['calcs_summary','Network Costs (demand)'        ] = 0
    data_gp[('calcs_summary','Market Profit (excl. network demand charges)')] = data_gp['calcs_summary'][Market_profit_list].sum(axis=1)
    print('   ***: Full Horizon Total Market Profit (excl. network demand charges) is '+ str(round(data_gp['calcs_summary'][Market_profit_list].sum(axis=1).sum(axis=0), 2))  )
    return data_gp

# ==================================================================================================================== 
def revenue_brkdwn(data_gp,lreg_rev,ppa_sim):
    # %% monthly revenue break down
    data_gp['calcs_rev','Energy Revenue'                 ] = data_gp['Rev']['energy_volumes_discharge']
    data_gp['calcs_rev','Raise Regulation Revenue'       ] = data_gp['Rev']['FCAS_rreg_volumes_discharge']
    data_gp['calcs_rev','Raise Regulation Energy Revenue'] = data_gp['FCAS_Energy_Revenue']['FCAS_rreg_volumes_discharge']
    data_gp['calcs_rev','1sec Raise Revenue'             ] = data_gp['Rev']['FCAS_r1sec_volumes_discharge']
    data_gp['calcs_rev','6sec Raise Revenue'             ] = data_gp['Rev']['FCAS_r6sec_volumes_discharge']
    data_gp['calcs_rev','60sec Raise Revenue'            ] = data_gp['Rev']['FCAS_r60sec_volumes_discharge']
    data_gp['calcs_rev','5min Raise Revenue'             ] = data_gp['Rev']['FCAS_r5min_volumes_discharge']
    data_gp['calcs_rev','Lower Regulation Revenue'       ] = data_gp['Rev'][lreg_rev].sum(axis=1)
    data_gp['calcs_rev','1sec Lower Revenue'             ] = data_gp['Rev']['FCAS_l1sec_volumes_charge']
    data_gp['calcs_rev','6sec Lower Revenue'             ] = data_gp['Rev']['FCAS_l6sec_volumes_charge']
    data_gp['calcs_rev','60sec Lower Revenue'            ] = data_gp['Rev']['FCAS_l60sec_volumes_charge']
    data_gp['calcs_rev','5min Lower Revenue'             ] = data_gp['Rev']['FCAS_l5min_volumes_charge']
    return data_gp

# ==================================================================================================================== 
def operation_calc(data_gp):
    # % operations
    data_gp['calcs_ops','Average State of Charge'               ] =  data_gp['Value']['soc_begin_period'            ]
    data_gp['calcs_ops','Reminaing capacity percentage'         ] =  data_gp['Value']['battery_current_cap']/data_gp['Value']['battery_current_cap'].iloc[0]
    data_gp['calcs_ops','State of Charge Max - End of Month'    ] =  data_gp['Value']['battery_current_cap'         ]
    data_gp['calcs_ops','Total Month Throughput'                ] =  data_gp['Throughput']['Period']
    data_gp['calcs_ops','Average Daily Throughput'              ] =  data_gp['calcs_ops','Total Month Throughput'   ]/ data_gp[('Datetime','Day')]
    data_gp['calcs_ops','Average Cycle Rate'                    ] =  data_gp['calcs_ops','Average Daily Throughput' ]/data_gp['Value','battery_current_cap']
    data_gp['calcs_ops','Average Energy Discharge when enabled' ] =  data_gp['Value','energy_volumes_discharge'     ]
    data_gp['calcs_ops','Average Energy Charge when enabled'    ] = -data_gp['Value','energy_volumes_charge'        ]
    data_gp['calcs_ops','Average Raise Regulation when enabled' ] =  data_gp['Value','FCAS_rreg_volumes_discharge'  ]
    data_gp['calcs_ops','Average Lower Regulation when enabled' ] = -data_gp['Value']['Total_Lreg']
    data_gp['calcs_ops','Average 1sec Raise when enabled'       ] =  data_gp['Value','FCAS_r1sec_volumes_discharge' ]
    data_gp['calcs_ops','Average 6sec Raise when enabled'       ] =  data_gp['Value','FCAS_r6sec_volumes_discharge' ]
    data_gp['calcs_ops','Average 60sec Raise when enabled'      ] =  data_gp['Value','FCAS_r60sec_volumes_discharge']
    data_gp['calcs_ops','Average 5min Raise when enabled'       ] =  data_gp['Value','FCAS_r5min_volumes_discharge' ]
    data_gp['calcs_ops','Average 1sec Lower when enabled'       ] = -data_gp['Value','FCAS_l1sec_volumes_charge'    ]
    data_gp['calcs_ops','Average 6sec Lower when enabled'       ] = -data_gp['Value','FCAS_l6sec_volumes_charge'    ]
    data_gp['calcs_ops','Average 60sec Lower when enabled'      ] = -data_gp['Value','FCAS_l60sec_volumes_charge'   ]
    data_gp['calcs_ops','Average 5min Lower when enabled'       ] = -data_gp['Value','FCAS_l5min_volumes_charge'    ]
    return data_gp

# ==================================================================================================================== 
def enabelement_calc(data_gp):
    # % Enablement
    data_gp['calcs_enbl','Energy Discharge' ] = data_gp['Count']['energy_volumes_discharge'     ] /  data_gp['Count']['periods']
    data_gp['calcs_enbl','Energy Charge'    ] = data_gp['Count']['energy_volumes_charge'        ] /  data_gp['Count']['periods']
    data_gp['calcs_enbl','Raise Regulation' ] = data_gp['Count']['FCAS_rreg_volumes_discharge'  ] /  data_gp['Count']['periods']
    data_gp['calcs_enbl','Lower Regulation' ] = data_gp['Count']['Total_Lreg'                   ] /  data_gp['Count']['periods']
    data_gp['calcs_enbl','1sec Raise'       ] = data_gp['Count']['FCAS_r1sec_volumes_discharge' ] /  data_gp['Count']['periods']
    data_gp['calcs_enbl','6sec Raise'       ] = data_gp['Count']['FCAS_r6sec_volumes_discharge' ] /  data_gp['Count']['periods']
    data_gp['calcs_enbl','60sec Raise'      ] = data_gp['Count']['FCAS_r60sec_volumes_discharge'] /  data_gp['Count']['periods']
    data_gp['calcs_enbl','5min Raise'       ] = data_gp['Count']['FCAS_r5min_volumes_discharge' ] /  data_gp['Count']['periods']
    data_gp['calcs_enbl','1sec Lower'       ] = data_gp['Count']['FCAS_l1sec_volumes_charge'    ] /  data_gp['Count']['periods']
    data_gp['calcs_enbl','6sec Lower'       ] = data_gp['Count']['FCAS_l6sec_volumes_charge'    ] /  data_gp['Count']['periods']
    data_gp['calcs_enbl','60sec Lower'      ] = data_gp['Count']['FCAS_l60sec_volumes_charge'   ] /  data_gp['Count']['periods']
    data_gp['calcs_enbl','5min Lower'       ] = data_gp['Count']['FCAS_l5min_volumes_charge'    ] /  data_gp['Count']['periods']
    return data_gp


# ==================================================================================================================== 
def throughput_calc(data_gp,BTM,resolution,Res_Load):
    #% Throughputs MWh
    data_gp['calcs_thrgh','Energy Discharge'] =  data_gp['Throughput','energy_volumes_discharge']
    data_gp['calcs_thrgh','Energy Charge'   ] = -data_gp['Throughput','energy_volumes_charge'   ]

    if BTM:
        data_gp['calcs_thrgh','BtM Charge from Curtailment (MWh)'           ] = -data_gp['Throughput','BTM_energy_vol_charge_curtail'] 
        data_gp['calcs_thrgh','BTM Charge from available generation (MWh)'  ] = -data_gp['Throughput']['BTM_energy_vol_charge_export_avail_gen']
        if Res_Load:
            data_gp['calcs_thrgh','BtM BESS to load (MWh)'  ] = data_gp['Throughput']['BTM_BESS_to_load'] 
            data_gp['calcs_thrgh','BtM Grid to load (MWh)'  ] = data_gp['Throughput']['BTM_Grid_to_load']
            data_gp['calcs_thrgh','BtM Solar to load (MWh)' ] = data_gp['Throughput']['BTM_solar_to_load']

    data_gp['calcs_thrgh','Raise Regulation'      ] =  data_gp['Throughput','FCAS_rreg_volumes_discharge']
    data_gp['calcs_thrgh','Total Lower Regulation'] = -data_gp['Throughput']['Total_Lreg'                           ]          
    data_gp['calcs_thrgh','BTM Lower Regulation'  ] = -data_gp['Throughput']['BTM_lreg_vol_charge_export_avail_gen' ]
    data_gp['calcs_thrgh','Grid Lower Regulation' ] = -data_gp['Throughput']['FCAS_lreg_volumes_charge'             ]
    
    data_gp['calcs_thrgh','1sec Raise'  ] =  data_gp['Throughput','FCAS_r1sec_volumes_discharge'  ]
    data_gp['calcs_thrgh','6sec Raise'  ] =  data_gp['Throughput','FCAS_r6sec_volumes_discharge'  ]
    data_gp['calcs_thrgh','60sec Raise' ] =  data_gp['Throughput','FCAS_r60sec_volumes_discharge' ]
    data_gp['calcs_thrgh','5min Raise'  ] =  data_gp['Throughput','FCAS_r5min_volumes_discharge'  ]
    data_gp['calcs_thrgh','1sec Lower'  ] = -data_gp['Throughput','FCAS_l1sec_volumes_charge'     ]
    data_gp['calcs_thrgh','6sec Lower'  ] = -data_gp['Throughput','FCAS_l6sec_volumes_charge'     ]
    data_gp['calcs_thrgh','60sec Lower' ] = -data_gp['Throughput','FCAS_l60sec_volumes_charge'    ]
    data_gp['calcs_thrgh','5min Lower'  ] = -data_gp['Throughput','FCAS_l5min_volumes_charge'     ]
    data_gp['calcs_thrgh','Dispatch'    ] =  data_gp['calcs_thrgh','Energy Discharge'] + data_gp['calcs_thrgh','Raise Regulation'      ] + data_gp['calcs_thrgh','6sec Raise'] + data_gp['calcs_thrgh','1sec Raise']  + data_gp['calcs_thrgh','60sec Raise'] + data_gp['calcs_thrgh','5min Raise']
    data_gp['calcs_thrgh','Charge'      ] =  data_gp['calcs_thrgh','Energy Charge'   ] + data_gp['calcs_thrgh','Total Lower Regulation'] + data_gp['calcs_thrgh','6sec Lower'] + data_gp['calcs_thrgh','1sec Lower']  + data_gp['calcs_thrgh','60sec Lower'] + data_gp['calcs_thrgh','5min Lower']
    
    if BTM:
        data_gp['calcs_thrgh','Charge'  ] = data_gp['calcs_thrgh','Charge'] + data_gp['calcs_thrgh','BtM Charge from Curtailment (MWh)'] + data_gp['calcs_thrgh','BTM Charge from available generation (MWh)']
    
    return data_gp

# ==================================================================================================================== 
def throughput_perc_calc(data_gp):
    #% Throughput percentage
    data_gp['calcs_thrgh_per','Energy Discharge'] =  data_gp['Throughput','energy_volumes_discharge'     ]/ data_gp['calcs_thrgh','Dispatch' ]
    data_gp['calcs_thrgh_per','Energy Charge'   ] = -data_gp['Throughput','energy_volumes_charge'        ]/ data_gp['calcs_thrgh','Charge'   ]
    data_gp['calcs_thrgh_per','Raise Regulation'] =  data_gp['Throughput','FCAS_rreg_volumes_discharge'  ]/ data_gp['calcs_thrgh','Dispatch' ]
    data_gp['calcs_thrgh_per','Lower Regulation'] = -data_gp['Throughput']['Total_Lreg'                  ]/ data_gp['calcs_thrgh','Charge'   ] 
    data_gp['calcs_thrgh_per','1sec Raise'      ] =  data_gp['Throughput','FCAS_r1sec_volumes_discharge' ]/ data_gp['calcs_thrgh','Dispatch' ]
    data_gp['calcs_thrgh_per','6sec Raise'      ] =  data_gp['Throughput','FCAS_r6sec_volumes_discharge' ]/ data_gp['calcs_thrgh','Dispatch' ]
    data_gp['calcs_thrgh_per','60sec Raise'     ] =  data_gp['Throughput','FCAS_r60sec_volumes_discharge']/ data_gp['calcs_thrgh','Dispatch' ]
    data_gp['calcs_thrgh_per','5min Raise'      ] =  data_gp['Throughput','FCAS_r5min_volumes_discharge' ]/ data_gp['calcs_thrgh','Dispatch' ]
    data_gp['calcs_thrgh_per','1sec Lower'      ] = -data_gp['Throughput','FCAS_l1sec_volumes_charge'    ]/ data_gp['calcs_thrgh','Charge'   ] 
    data_gp['calcs_thrgh_per','6sec Lower'      ] = -data_gp['Throughput','FCAS_l6sec_volumes_charge'    ]/ data_gp['calcs_thrgh','Charge'   ] 
    data_gp['calcs_thrgh_per','60sec Lower'     ] = -data_gp['Throughput','FCAS_l60sec_volumes_charge'   ]/ data_gp['calcs_thrgh','Charge'   ] 
    data_gp['calcs_thrgh_per','5min Lower'      ] = -data_gp['Throughput','FCAS_l5min_volumes_charge'    ]/ data_gp['calcs_thrgh','Charge'   ] 
    return data_gp

# ==================================================================================================================== 
def BTM_calc(data_gp,resolution,Res_Load):
    #% BTM charging
    data_gp['calcs_BTM','Gross Solar Generation (MWh)'                  ] =  data_gp['Curtail'      ]['net_co_locate_gen_avail_export'           ] * resolution + data_gp['Curtail','total_curt_avail'] * resolution
    data_gp['calcs_BTM','Revenue Actual Generation Exported'            ] =  data_gp['Rev'          ]['BTM_co_loc_gen_energy_export'             ] 
    data_gp['calcs_BTM','Cost BtM Charge from available Generation'     ] =  data_gp['Rev'          ]['BTM_energy_vol_charge_export_avail_gen'   ] 
    data_gp['calcs_BTM','LREG Charge BTM (MWh)'                         ] =  data_gp['Throughput'   ]['BTM_lreg_vol_charge_export_avail_gen'     ]
    data_gp['calcs_BTM','Charge from available generation (MWh)'        ] = -data_gp['Throughput'   ]['BTM_energy_vol_charge_export_avail_gen'   ]
    data_gp['calcs_BTM','Coloc Generation Exported (MWh)'               ] =  data_gp['Value'        ]['BTM_co_loc_gen_energy_export'             ] * resolution 
    data_gp['calcs_BTM','Actual Solar Generation Exported(MWh)'         ] =  data_gp['calcs_BTM' ,'Coloc Generation Exported (MWh)'              ] +  data_gp['calcs_BTM','LREG Charge BTM (MWh)'       ]
    data_gp['calcs_BTM','percent of Solar Generation used for charging' ] =  data_gp['calcs_BTM' ,'Charge from available generation (MWh)'       ] /  data_gp['calcs_BTM','Gross Solar Generation (MWh)']
    data_gp['calcs_BTM','percent of Solar Generation exported'          ] =  data_gp['calcs_BTM' ,'Actual Solar Generation Exported(MWh)'        ] /  data_gp['calcs_BTM','Gross Solar Generation (MWh)']
    data_gp['calcs_BTM','Curtailment volume available for charging'     ] =  data_gp['Curtail'   ,'total_curt_avail'             ] * resolution
    data_gp['calcs_BTM','Unavoidable curtailment'                       ] =  data_gp['Curtail'   ,'total_unaviod_curt'           ] * resolution
    data_gp['calcs_BTM','BtM Charge from Curtailment (MWh)'             ] = -data_gp['Throughput','BTM_energy_vol_charge_curtail'] 
    data_gp['calcs_BTM','LREG Charge BTM ($)'                           ] =  data_gp[('FCAS_Energy_Revenue','BTM_lreg_vol_charge_export_avail_gen')]
    data_gp['calcs_BTM','Curtailment for BESS FCAS provision ($)'       ] = -data_gp[('Cost'     ,'BESS_Con_Res_Curtail')]
    data_gp['calcs_BTM','Curtailment for BESS FCAS provision (MWh)'     ] =  data_gp[('Value'    ,'BESS_Con_Res_Curtail')]
    #%% Residential load 
    if Res_Load:
        data_gp['calcs_BTM','Gross Residential Load (MWh)'                         ] = data_gp['Residential_load']['net_co_locate_load_avail_export'] * resolution 
        data_gp['calcs_BTM','BTM_BESS_to_load (MWh)'                              ] = data_gp['Throughput','BTM_BESS_to_load'   ] 
        data_gp['calcs_BTM','BTM_Grid_to_load (MWh)'                              ] = -data_gp[('Throughput','BTM_Grid_to_load') ]
        data_gp['calcs_BTM','BTM_solar_to_load (MWh)'                             ] = data_gp[('Throughput','BTM_solar_to_load')]
        data_gp['calcs_BTM','percent of BESS used for load'                       ] = data_gp[('Throughput','BTM_BESS_to_load') ] /  data_gp['calcs_BTM','Gross Residential Load (MWh)'  ]
        data_gp['calcs_BTM','percent of Grid  used for load'                      ] = -data_gp[('Throughput','BTM_Grid_to_load') ] /  data_gp['calcs_BTM','Gross Residential Load (MWh)'  ]
        data_gp['calcs_BTM','percent of Solar Generation used for load'           ] = data_gp[('Throughput','BTM_solar_to_load')] /  data_gp['calcs_BTM','Gross Residential Load (MWh)'  ]
        data_gp['calcs_BTM','Average Price of Solar when used for load ($/MWh) (Not implemented)'                ] = data_gp[('Throughput','BTM_solar_to_load') ] /  data_gp['calcs_BTM','Gross Residential Load (MWh)'  ]
        data_gp['calcs_BTM','Average Price of BESS when used for load ($/MWh) (Not implemented)'                 ] = data_gp[('Throughput','BTM_BESS_to_load') ] /  data_gp['calcs_BTM','Gross Residential Load (MWh)'  ]
        data_gp['calcs_BTM','Average Price of Grid  when used for load ($/MWh) (Not implemented)'                ] = -data_gp[('Throughput','BTM_Grid_to_load')] /  data_gp['calcs_BTM','Gross Residential Load (MWh)'  ]
    return data_gp

# ==================================================================================================================== 
def PPA_calc(data_gp):
    data_gp['calcs_PPA','PPA Price ($/MWh)'                     ] = data_gp[ 'Price']['PPA']
    data_gp['calcs_PPA','PPA offtake (MWh)'                     ] = data_gp[ 'Throughput', 'PPA_offtake'        ]
    data_gp['calcs_PPA','PPA Load Satisfied'                    ] = data_gp[('Throughput', 'PPA Load Satisfied')]
    data_gp['calcs_PPA','PPA Load Exceed'                       ] = data_gp[ 'Throughput', 'PPA_Load_Exceed'    ]
    data_gp['calcs_PPA','PPA  Load Satisfied - Spot Revenue($)' ] = data_gp['Rev']['PPA Load Satisfied Spot'    ]
    # data_gp['calcs_PPA','PPA Revenue ($)'] = data_gp['Rev']['PPA offtake Total']
    data_gp['calcs_PPA','Cost of PPA Load Exceed'               ] = -data_gp['Cost','PPA Load Exceed']
    data_gp['calcs_PPA','PPA CFD Adjustment($)'                 ] =  data_gp[('Rev', 'PPA Revenue Load Satisfied') ] - data_gp['calcs_PPA','PPA  Load Satisfied - Spot Revenue($)']
    data_gp['calcs_PPA','PPA Total Adjustment($)'               ] =  data_gp['calcs_PPA','Cost of PPA Load Exceed' ] + data_gp['calcs_PPA','PPA CFD Adjustment($)'                ]
    # data_gp['calcs_PPA','PPA offtake Solar (MWh)'] = data_gp[('Throughput' , 'PPA offtake Solar')] 
    return data_gp

# ==================================================================================================================== 
def rev_per_MWh(data_gp,data_sum):
    #% $/MW or $/MWh
    data_gp['calcs_RevperCap', 'Energy Revenue ($/MWh)'                  ] =  data_gp['calcs_rev'    ,'Energy Revenue'                   ] / data_sum['Throughput'  ,'energy_volumes_discharge'     ]
    data_gp['calcs_RevperCap', 'Energy Cost ($/MWh)'                     ] = -data_gp['calcs_summary','Energy Costs'                     ] / data_sum['Throughput'  ,'energy_volumes_charge'        ]
    data_gp['calcs_RevperCap', 'Raise Regulation Revenue ($/MW)'         ] =  data_gp['calcs_rev'    ,'Raise Regulation Revenue'         ] / data_sum['Value'       ,'FCAS_rreg_volumes_discharge'  ]
    data_gp['calcs_RevperCap', 'Raise Regulation Energy Revenue ($/MWh)' ] =  data_gp['calcs_rev'    ,'Raise Regulation Energy Revenue'  ] / data_sum['Throughput'  ,'FCAS_rreg_volumes_discharge'  ]
    data_gp['calcs_RevperCap', '1sec Raise Revenue ($/MW)'               ] =  data_gp['calcs_rev'    ,'1sec Raise Revenue'               ] / data_sum['Value'       ,'FCAS_r1sec_volumes_discharge' ]
    data_gp['calcs_RevperCap', '6sec Raise Revenue ($/MW)'               ] =  data_gp['calcs_rev'    ,'6sec Raise Revenue'               ] / data_sum['Value'       ,'FCAS_r6sec_volumes_discharge' ]
    data_gp['calcs_RevperCap', '60sec Raise Revenue ($/MW)'              ] =  data_gp['calcs_rev'    ,'60sec Raise Revenue'              ] / data_sum['Value'       ,'FCAS_r60sec_volumes_discharge']
    data_gp['calcs_RevperCap', '5min Raise Revenue ($/MW)'               ] =  data_gp['calcs_rev'    ,'5min Raise Revenue'               ] / data_sum['Value'       ,'FCAS_r5min_volumes_discharge' ]
    data_gp['calcs_RevperCap', 'Lower Regulation Revenue ($/MW)'         ] = -data_gp['calcs_rev'    ,'Lower Regulation Revenue'         ] / data_sum['Value']['Total_Lreg']
    data_gp['calcs_RevperCap', 'Lower Regulation Energy Cost ($/MWh)'    ] =  data_gp['calcs_summary','Lreg Energy Costs'                ] / data_sum['Throughput'  ,'FCAS_rreg_volumes_discharge'  ]
# if BTM:
#     data_gp['calcs_RevperCap','BTM Charge From Avail Export ($/MWh)']   = data_gp['Rev']['BTM_energy_vol_charge_export_avail_gen'] /  data_sum['Throughput' , 'BTM_lreg_vol_charge_export_avail_gen']
#     data_gp['calcs_RevperCap','BTM Lower Regulation Revenue ($/MW)'] = data_gp['Rev']['FCAS_lreg_volumes_charge']/ data_sum['Value' , 'BTM_lreg_vol_charge_export_avail_gen']
    data_gp['calcs_RevperCap','1sec Lower Revenue ($/MW)'                ] = -data_gp['calcs_rev','1sec Lower Revenue'                   ] / data_sum['Value','FCAS_l1sec_volumes_charge'           ]
    data_gp['calcs_RevperCap','6sec Lower Revenue ($/MW)'                ] = -data_gp['calcs_rev','6sec Lower Revenue'                   ] / data_sum['Value','FCAS_l6sec_volumes_charge'           ]
    data_gp['calcs_RevperCap','60sec Lower Revenue ($/MW)'               ] = -data_gp['calcs_rev','60sec Lower Revenue'                  ] / data_sum['Value','FCAS_l60sec_volumes_charge'          ]
    data_gp['calcs_RevperCap','5min Lower Revenue ($/MW)'                ] = -data_gp['calcs_rev','5min Lower Revenue'                   ] / data_sum['Value','FCAS_l5min_volumes_charge'           ]
    return data_gp

# ==================================================================================================================== 
def DLF_MLF_rev(data_gp):
    #% DLF MLF revenue loss
    data_gp['calcs_DLF_MLF','DLF Energy Revenue Loss'                    ] = -data_gp['calcs_rev'    ,'Energy Revenue'                   ] / (data_gp['DLF']['DLF_BESS'] * data_gp['MLF']['MLF_generation']) * (1- data_gp['DLF']['DLF_BESS'      ])
    data_gp['calcs_DLF_MLF','MLF Energy Revenue Loss'                    ] = -data_gp['calcs_rev'    ,'Energy Revenue'                   ] / (data_gp['DLF']['DLF_BESS'] * data_gp['MLF']['MLF_generation']) * (1- data_gp['MLF']['MLF_generation'])
    data_gp['calcs_DLF_MLF','DLF Raise Regulation Energy Revenue Loss'   ] = -data_gp['calcs_rev'    ,'Raise Regulation Energy Revenue'  ] / (data_gp['DLF']['DLF_BESS'] * data_gp['MLF']['MLF_generation']) * (1- data_gp['DLF']['DLF_BESS'      ])
    data_gp['calcs_DLF_MLF','MLF Raise Regulation Energy Revenue Loss'   ] = -data_gp['calcs_rev'    ,'Raise Regulation Energy Revenue'  ] / (data_gp['DLF']['DLF_BESS'] * data_gp['MLF']['MLF_generation']) * (1- data_gp['MLF']['MLF_generation'])
    data_gp['calcs_DLF_MLF','DLF Energy Cost Benefit'                    ] = -data_gp['calcs_summary','Energy Costs'                     ] / (data_gp['DLF']['DLF_BESS'] * data_gp['MLF']['MLF_load'      ]) * (1- data_gp['DLF']['DLF_BESS'      ])
    data_gp['calcs_DLF_MLF','MLF Energy Cost Benefit'                    ] = -data_gp['calcs_summary','Energy Costs'                     ] / (data_gp['DLF']['DLF_BESS'] * data_gp['MLF']['MLF_load'      ]) * (1- data_gp['MLF']['MLF_load'      ])
    data_gp['calcs_DLF_MLF','DLF Lower Regulation Energy Cost Benefit'   ] = -data_gp['calcs_summary','Lreg Energy Costs'                ] / (data_gp['DLF']['DLF_BESS'] * data_gp['MLF']['MLF_load'      ]) * (1- data_gp['DLF']['DLF_BESS'      ])
    data_gp['calcs_DLF_MLF','MLF Lower Regulation Energy Cost Benefit'   ] = -data_gp['calcs_summary','Lreg Energy Costs'                ] / (data_gp['DLF']['DLF_BESS'] * data_gp['MLF']['MLF_load'      ]) * (1- data_gp['MLF']['MLF_load'      ])
    return data_gp

# ==================================================================================================================== 
def aemo_cost(aemo_fees,region):
    aemo_fees           =  aemo_fees.loc[aemo_fees['State'] == region]
    aemo_fees           =  aemo_fees.pivot(index='State',columns='Fee or market charge',values= 'Rate' )
    aemo_customer_cost  = (aemo_fees['Market Customer Ancillary Services' ] + aemo_fees['AEMO General Fees unallocated'] + aemo_fees['AEMO Allocated Fees']).iloc[0]
    aemo_generator_cost = (aemo_fees['Market Generator Ancillary Services'] ).iloc[0]
    aemo_daily_cost     =  aemo_fees['AEMO Allocated Fees daily'          ]  .iloc[0]

    aemo_costs_dict = {'aemo_gen'       :aemo_generator_cost,
                       'aemo_customer'  :aemo_customer_cost , 
                       'aemo_daily'     :aemo_daily_cost    }
    return aemo_costs_dict

# ==================================================================================================================== 
def merge_curtail(data,curtail):
    if not data.columns.isin([('Datetime','Date')]).sum():
        data[('Datetime' ,'Date')] = data[('Datetime','')].dt.date
    x         = data[[('Datetime','Date'),('Period','')]]
    x.columns = ['Date' , 'Period']
    x['Date'] = pd.to_datetime(x['Date'])
    x = x.merge(curtail,on=['Date','Period'],how='left') 
    for s in x.columns[3:]:
        data[('Curtail',s)] = x[s]
    return data

def merge_residential(data,residential_load):
    if not data.columns.isin([('Datetime','Date')]).sum():
        data[('Datetime' ,'Date')] = data[('Datetime','')].dt.date
    x         = data[[('Datetime','Date'),('Period','')]]
    x.columns = ['Date' , 'Period']
    x['Date'] = pd.to_datetime(x['Date'])
    x = x.merge(residential_load,on=['Date','Period'],how='left') 
    for s in x.columns[3:]:
        data[('Residential_load',s)] = x[s]
    return data
# ==================================================================================================================== 
def merge_DLF_MLF(data,MLF,DLF):
    data.columns = [f"{level}**{col}" if level else col for level, col in data.columns] # Flatten for merge
    x = data.merge(MLF,left_on = [('Datetime**Year')],right_on= [('Year')],how='left')
   
    cols = MLF.columns[1:]
    for c in cols:
        data['MLF**' + c] = x[c]
    
    x = data.merge(DLF,left_on = [('Datetime**Year')],right_on= [('Year')],how='left')    

    data['DLF**DLF_BESS'] = x['DLF_BESS']

    if DLF.columns.isin(['DLF_cogen']).sum():
        data['DLF**DLF_cogen'] = x['DLF_cogen']

    # Create MultiIndex DataFrame
    # Initialize two arrays to store values before and after '_'
    before_underscore_array = []
    after_underscore_array = []

    # Iterate over each string in the array
    for s in data.columns:
        # Split the string into two parts based on '_'
        parts = s.split('**', 1)
        # Add the parts before '_' to the array
        before_underscore_array.append(parts[0])
        # Add the parts after '_' to the array
        after_underscore_array.append(parts[1]) 

    # Create MultiIndex array
    multi_index = pd.MultiIndex.from_arrays([before_underscore_array, after_underscore_array], names=['One', 'Two'])

    # Create DataFrame with MultiIndex
    data = pd.DataFrame(data.values, columns=multi_index)
    
    return data

# ==================================================================================================================== 
def merge_DNSP(data,DNSP):
    data.columns = [f"{level}**{col}" if level else col for level, col in data.columns] # Flatten for merge

    x    = data.merge(DNSP, left_on = [('Datetime**Year')], right_on= 'Year',how='left')
    cols = ['Fixed ($/day)','Capacity Charge ($/kVA/month)','Demand Peak ($/kVA/month)' ]
    for c in cols:
        data['DNSP**'+c] = x[c]

    # Create MultiIndex DataFrame
    # Initialize two arrays to store values before and after '_'
    before_underscore_array = []
    after_underscore_array = []

    # Iterate over each string in the array
    for s in data.columns:
        # Split the string into two parts based on '_'
        parts = s.split('**', 1)
        # Add the parts before '_' to the array
        before_underscore_array.append(parts[0])
        # Add the parts after '_' to the array
        after_underscore_array.append(parts[1]) 

    # Create MultiIndex array
    multi_index = pd.MultiIndex.from_arrays([before_underscore_array, after_underscore_array], names=['One', 'Two'])

    # Create DataFrame with MultiIndex
    data = pd.DataFrame(data.values, columns=multi_index)    
    return data

# ==================================================================================================================== 
def merge_LGC(data,LGC):
    data.columns = [f"{level}**{col}" if level else col for level, col in data.columns] # Flatten for merge
    x = data.merge(LGC, left_on=[('Datetime**Year')], right_on='Year', how='left')
    data['LGC**Price'] = x   ['LGC price'  ]
    data['LGC**Price'] = data['LGC**Price'].astype(float)
    
      # Create MultiIndex DataFrame
    # Initialize two arrays to store values before and after '_'
    before_underscore_array = []
    after_underscore_array = []

    # Iterate over each string in the array
    for s in data.columns:
        # Split the string into two parts based on '_'
        parts = s.split('**', 1)
        # Add the parts before '_' to the array
        before_underscore_array.append(parts[0])
        
        # Add the parts after '_' to the array
        after_underscore_array.append(parts[1]) 

    # Create MultiIndex array
    multi_index = pd.MultiIndex.from_arrays([before_underscore_array, after_underscore_array], names=['One', 'Two'])

    # Create DataFrame with MultiIndex
    data = pd.DataFrame(data.values, columns=multi_index)
    
    return data

# ==================================================================================================================== 
def merge_PPA(data,PPA):
    data.columns = [f"{level}_{col}" if level else col for level, col in data.columns] # Flatten for merge
    x = data.merge(PPA, left_on=[('Datetime','Year')], right_on='Year', how='left')
    data['Price','PPA'] = x['PPA_Price']
    data['Price','PPA'] = data['Price','PPA'].astype(float)
    
  # Create MultiIndex DataFrame
    # Initialize two arrays to store values before and after '_'
    before_underscore_array = []
    after_underscore_array = []

    # Iterate over each string in the array
    for s in data.columns:
        # Split the string into two parts based on '_'
        parts = s.split('_')
        
        # Add the parts before '_' to the array
        before_underscore_array.append(parts[0])
        
        # Add the parts after '_' to the array
        after_underscore_array.append(parts[1]) 

    # Create MultiIndex array
    multi_index = pd.MultiIndex.from_arrays([before_underscore_array, after_underscore_array], names=['One', 'Two'])

    # Create DataFrame with MultiIndex
    data = pd.DataFrame(data.values, columns=multi_index)

    return data

# ==================================================================================================================== 
def monthly_data(data,svc_list):
    # List of specific columns to exclude from conversion
    exclude_columns = [('Datetime','Year'), ('Datetime','Month'), ('Datetime',''), ('Period','') ]
    # Identify columns to convert to numeric (excluding those in exclude_columns)
    numeric_columns = data.columns.difference(exclude_columns)
    # Convert identified columns to numeric
    data[numeric_columns] = data[numeric_columns].apply(pd.to_numeric, errors='coerce')
 
    data_sum            =  data                 .groupby([('Datetime','Year'), ('Datetime','Month')],as_index=False).sum(numeric_only=True)
    data_max            =  data                   .groupby([('Datetime','Year'), ('Datetime','Month')],as_index=False).max(numeric_only=True)
    data_mean           =  data                   .groupby([('Datetime','Year'), ('Datetime','Month')],as_index=False).mean(numeric_only=True)
    data_mean_nonzeros  = (data.replace(0,np.nan)).groupby([('Datetime','Year'), ('Datetime','Month')],as_index=False).mean(numeric_only=True)
    data_min            =  data                   .groupby([('Datetime','Year'), ('Datetime','Month')],as_index=False).min(numeric_only=True)
    data_count          = (data.replace(0,np.nan)).groupby([('Datetime','Year'), ('Datetime','Month')],as_index=False).count()
    data_gp             =  data_sum               .copy()
    #number of days per month for fix network cost
    data_gp[('Datetime','Day'               )] = data_max [('Datetime','Day'                )]
    # avg SOC
    data_gp[('Value','soc_begin_period'     )] = data_mean[('Value'   ,'soc_begin_period'   )]
    # max battery cap end of each month
    data_gp[('Value','battery_current_cap'  )] = data_min [('Value'   ,'battery_current_cap')]
    # price
    data_gp[('Price'                        )] = data_mean[('Price'                         )]

    # avg enablement
    for s in svc_list: 
          data_gp [('Value',s )] = data_mean_nonzeros['Value', s]
          data_gp ['Count' ,s  ] = data_count        ['Value', s]
    data_gp['Count','periods'] = data_count        ['Datetime','Day']  #max count
    data_gp['MLF'            ] = data_mean         ['MLF'           ]  #mean MLF
    data_gp['DLF'            ] = data_mean         ['DLF'           ]  #mean DLF
    data_gp['DNSP'           ] = data_mean         ['DNSP'          ]    
    return data_gp, data_sum

# ==================================================================================================================== 
def demand_charge_monthly(df   , tariff  , peak_start=15 , peak_end=39 , peak_numbers=1 ,  pf = 0.93):
    """ tariff is $/kVA/month"""
    df[('Value','Total_charge')] = df[('Value','energy_volumes_charge')] #+ df[('Value','FCAS_lreg_volumes_charge')]
    df_dem_peak                         = df.loc[(df[('Period','Unnamed: 2_level_1')] >= peak_start) & (df[('Period','Unnamed: 2_level_1')] <=peak_end) ] 
    df_dem_peak[('Value','MVA')]        = np.sqrt(((1- pf)*df[('Value','Total_charge')])**2 +(df[('Value','Total_charge')])**2 )
    dem_peak                            = df_dem_peak.groupby([('Datetime','Year'), ('Datetime','Month')])[[('Value','MVA')]].apply(lambda x: x.nlargest(1,('Value','MVA')).mean())
    dem_peak[('Cost','Demand_charge')]  = dem_peak[('Value','MVA')] * tariff * 1000 #kVA to MVA
    return dem_peak[('Cost','Demand_charge')]

# ==================================================================================================================== 
def export_summary(df, BTM ,Res_Load, ppa_sim, out_dir,file_name,Solve_status_dict):
    files_path = os.path.join(out_dir,'output_'+ file_name+ '.xlsx')
    out_file = files_path
    if BTM:
        cols = ['Datetime','calcs_summary','calcs_rev','calcs_ops','calcs_enbl', 'calcs_thrgh','calcs_thrgh_per' ,'calcs_BTM','calcs_RevperCap','calcs_DLF_MLF']
        if ppa_sim:
            cols = cols + ['calcs_PPA']
    else:   
        cols = ['Datetime','calcs_summary','calcs_rev','calcs_ops','calcs_enbl', 'calcs_thrgh','calcs_thrgh_per' ,'calcs_RevperCap','calcs_DLF_MLF']

    df_out      = df[cols]
    df_out      = df_out.drop([('Datetime','Day' )], axis=1)
    df_annaul   = df.groupby ([('Datetime','Year')], as_index=False).sum()
    writer = pd.ExcelWriter(out_file)

    # date
    row_gaps       = 3
    Date           = df_out['Datetime']
    Date           .T.to_excel(writer, sheet_name='Dataset', header=False , startcol=0,startrow=1)

    # summary
    summary        = df_out['calcs_summary']
    row_ind_summary= 2+Date.shape[1]
    summary        .T.to_excel(writer, sheet_name='Dataset' ,header=False, startcol=0, startrow=row_ind_summary)

    # revenue break down
    rev            = df_out['calcs_rev']
    row_ind_rev    = row_ind_summary + summary.shape[1] + row_gaps
    rev            .T.to_excel(writer, sheet_name='Dataset' ,header=False, startcol=0, startrow=row_ind_rev)

    # operation 
    ops            = df_out['calcs_ops']
    cols           = ops.columns
    cols_order     = cols[0:3].append(cols[5:]).append(cols[4:2:-1])
    ops            = ops[cols_order]
    row_ind_ops    = row_ind_rev + rev.shape[1] + row_gaps
    ops            .T.to_excel(writer, sheet_name='Dataset', header=False, startcol=0, startrow=row_ind_ops)

    #enablement
    enbl           = df_out['calcs_enbl']
    row_ind_enbl   = row_ind_ops + ops.shape[1] + row_gaps
    enbl           .T.to_excel(writer, sheet_name='Dataset', header=False, startcol=0, startrow=row_ind_enbl)

    #throughput
    thr            = df_out['calcs_thrgh']
    row_ind_thr    = row_ind_enbl + enbl.shape[1] + row_gaps
    thr            .T.to_excel(writer, sheet_name='Dataset', header=False, startcol=0, startrow=row_ind_thr)

    #throughput percentages
    thrper         = df_out['calcs_thrgh_per']
    row_ind_thrper = row_ind_thr + thr.shape[1] + row_gaps
    thrper         .T.to_excel(writer, sheet_name='Dataset', header=False, startcol=0, startrow=row_ind_thrper)

    # Revenue per MW or MWh
    rev_per        = df_out['calcs_RevperCap']
    row_ind_revper = row_ind_thrper + thrper.shape[1] + row_gaps
    rev_per        .T.to_excel(writer, sheet_name='Dataset', header=False, startcol=0, startrow=row_ind_revper)

    # DLF/MLF costs
    rev_dmlf       = df_out['calcs_DLF_MLF']
    row_ind_dmlf   = row_ind_revper + rev_per.shape[1] + row_gaps
    rev_dmlf       .T.to_excel(writer, sheet_name='Dataset', header=False, startcol=0, startrow=row_ind_dmlf)

    if BTM:
        #BTM monthly
        rev_btm     = df_out['calcs_BTM']
        row_ind_btm = row_ind_dmlf + rev_dmlf.shape[1] + row_gaps
        rev_btm     .T.to_excel(writer, sheet_name='Dataset', header=False, startcol=0, startrow=row_ind_btm)
        row_gaps    = 3

        #date
        Date       = df_annaul['Datetime']
        Date       .T.to_excel(writer, sheet_name='BTM', header=False , startcol=0, startrow=1)    
        # writing BTM calcs

        row_BTM   = 2+Date.shape[1]
        BTM_calcs = df_annaul['calcs_BTM']
        BTM_calcs .T.to_excel(writer, sheet_name='BTM', header=False, startcol=0, startrow=row_BTM)
        if ppa_sim:
            #PPA monthly
            rev_ppa     = df_out['calcs_PPA']
            row_ind_ppa = row_ind_btm + rev_btm.shape[1] + row_gaps
            rev_ppa     .T.to_excel(writer, sheet_name='Dataset', header=False, startcol=0,startrow=row_ind_ppa)
            row_gaps    = 3

            #date
            Date       = df_annaul['Datetime']
            Date       .T.to_excel(writer,sheet_name='PPA',header=False , startcol=0,startrow=1)   
 
            # writing PPA calcs
            row_PPA = 2+Date.shape[1]
            PPA_calcs = df_annaul['calcs_PPA']
            PPA_calcs.T.to_excel(writer,sheet_name='PPA' ,header=False, startcol=0,startrow=row_PPA)

    #writer.save()
    writer.close()

    # writing title of each section of Dataset tab
    wb = pyxl.load_workbook(filename=out_file)
    ws = wb.worksheets[0]
    ws['A'+str(row_ind_summary  )] = 'Summary'
    ws['A'+str(row_ind_rev      )] = 'Market Revenue Breakdown'
    ws['A'+str(row_ind_ops      )] = 'Operations'
    ws['A'+str(row_ind_enbl     )] = 'Enablement'
    ws['A'+str(row_ind_thr      )] = 'Throughputs'
    ws['A'+str(row_ind_thrper   )] = 'Throughputs Percentage'
    ws['A'+str(row_ind_revper   )] = 'Revenue per MW or MWh'
    ws['A'+str(row_ind_dmlf     )] = 'DLF/MLF costs or benefits'
    if BTM:
        ws['A'+str(row_ind_btm      )] = 'BTM'
        if ppa_sim:
            ws['A'+str(row_ind_ppa      )] = 'PPA'
    wb.save(out_file)
    
    
    output_to_proforma(out_dir,file_name,BTM,Res_Load,ppa_sim,Solve_status_dict)

def output_to_proforma(out_dir,file_name,BTM,Res_Load,ppa_sim,Solve_status_dict): 

    
    #Pulls files from input folder and places them in the outputs 
    #Therefore the exact inputs that were used to produce the results can be accessed with any results.
    files_path  = Path().absolute().parents[0]
    inputs_path = os.path.join(files_path, 'Inputs')
    Sim_Run_Csv = pd.read_csv(inputs_path + "\\SIM Run.csv")
    Scenario_Manager_excel = pd.read_excel(inputs_path + "\\ScenarioManager.xlsx")

    Sim_Run_Csv.to_csv(out_dir + "\\SIM Run.csv")
    Scenario_Manager_excel.to_excel(out_dir + "\\ScenarioManager.xlsx")
    
    
        #     x="The problem is infeasible."


        #     x="The problem is unbounded."


        #     x="The solver did not find an optimal solution."
    
    if 'The problem is infeasible.' in Solve_status_dict.values() or 'The problem is unbounded.' in Solve_status_dict.values() or'The solver did not find an optimal solution.' in Solve_status_dict.values():

        Solve_status_df=pd.DataFrame.from_dict(Solve_status_dict, orient='index')
        Solve_status_df.to_csv(r"{}\\Solve Status log, error found.csv".format(out_dir))
        
        
    
    
    if BTM:
        if Res_Load:
            files_path = os.path.join(out_dir,'output_'+ file_name+ '.xlsx')
            proforma_output_path=os.path.join(out_dir,'Proforma' + file_name+ ' Res load.xlsx')
            Proforma_path="Template_Proforma/Proforma BESS (Scenario xx - Region xxMW-xxxMWh) - Residential load Wholesale BtM.xlsx"
            output_dataset=pd.read_excel(files_path, 'Dataset')
            df_proforma=output_dataset.iloc[3:,1:]    
            df_proforma_year_month=output_dataset.iloc[0:2,1:] 
            wb_name = Proforma_path
            sheet_name = 'SIM Dataset'
            df_mapping = {"C2":df_proforma_year_month,"C8": df_proforma}
            Write_to_workbook(sheet_name,df_mapping,wb_name,proforma_output_path)
            
        elif ppa_sim:
          
            files_path = os.path.join(out_dir,'output_'+ file_name+ '.xlsx')
            proforma_output_path=os.path.join(out_dir,'Proforma' + file_name+ ' PPA.xlsx')
            Proforma_path="Template_Proforma/Proforma BESS (Scenario xx - Region xMW-xxMWh) - PPA + BESS + Solar BtM.xlsx"
            output_dataset=pd.read_excel(files_path, 'Dataset')
            df_proforma=output_dataset.iloc[3:101,1:]    
            df_proforma_year_month=output_dataset.iloc[0:2,1:] 
            df_PPA_proforma=output_dataset.iloc[150:,1:]    
            wb_name = Proforma_path
            sheet_name = 'Dataset'
            df_mapping = {"C2":df_proforma_year_month,"C8": df_proforma,"C152":df_PPA_proforma}
            Write_to_workbook(sheet_name,df_mapping,wb_name,proforma_output_path)
        
        else:   
            files_path = os.path.join(out_dir,'output_'+ file_name+ '.xlsx')
            proforma_output_path=os.path.join(out_dir,'Proforma' + file_name+ ' BtM.xlsx')
            Proforma_path="Template_Proforma/Proforma BESS (Scenario xx - Region xxMW-xxxMWh) - BESS + Solar co-located BtM.xlsx"
            output_dataset=pd.read_excel(files_path, 'Dataset')
            df_proforma=output_dataset.iloc[3:101,1:]    
            df_proforma_year_month=output_dataset.iloc[0:2,1:]  
            
            wb_name = Proforma_path
            sheet_name = 'Dataset'
            df_mapping = {"C2":df_proforma_year_month,"C8": df_proforma}
            Write_to_workbook(sheet_name,df_mapping,wb_name,proforma_output_path)

    else:
        files_path = os.path.join(out_dir,'output_'+ file_name+ '.xlsx')
        proforma_output_path=os.path.join(out_dir,'Proforma' + file_name+ ' Standalone BESS.xlsx')
        Proforma_path="Template_Proforma/Proforma (Scenario xx - Region xxMW-xxxMWh) - StandaloneBESS+FinModel.xlsx"
        output_dataset=pd.read_excel(files_path, 'Dataset')
        df_proforma=output_dataset.iloc[3:99,1:]    
        df_proforma_year_month=output_dataset.iloc[0:2,1:]  
    
        wb_name = Proforma_path
        sheet_name = 'SIM Dataset'
        df_mapping = {"C2":df_proforma_year_month,"C8": df_proforma}
        #df_mapping = {"C7": df_proforma}
        Write_to_workbook(sheet_name,df_mapping,wb_name,proforma_output_path)
    

def Write_to_workbook(sheet_name,df_mapping,wb_name,proforma_output_path):    
        # Open Excel in background
        with xlwings.App(visible=False) as app:
            wb = app.books.open(wb_name)
            # Add sheet if it does not exist
            current_sheets = [sheet.name for sheet in wb.sheets]
            if sheet_name not in current_sheets:
                wb.sheets.add(sheet_name)
            # Write dataframe to cell range
            for cell_target, df in df_mapping.items():
                wb.sheets(sheet_name).range(cell_target).options(pd.DataFrame, index=False,header=False).value = df
            output_wb=wb
            output_wb.save(proforma_output_path)
            #wb.save()
        
